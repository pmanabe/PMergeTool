import glob
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import urllib.parse

templateHeaders = {
    'Full Name': [''],
    'Description': [''],
    'User License': [''],
    'Apex Class Access': ['Class'],
    'Application Access': ['Application'],
    'Field Level Security (FLS)': ['Object', 'Field'],
    'Login IP Ranges': ['Description', 'Start', 'End'],
    'Object Level Security (OLS)': ['Object'],
    'Record Type Creation': ['Object', 'Record Type'],
    'Tab Access': ['Tab'],
    'User Permission Access': ['User Permission'],
    'Visualforce Page Access': ['Page'],
    'Custom Metadata Type Access': ['Access'],
    'Custom Setting Access': ['Access'],
    'License': [''],
    'Page Layout Definitions': ['Object', 'Layout']
}

templateStartRowColumn = {
    'Full Name': [1,1],
    'Description': [1,1],
    'User License': [1,1],
    'Apex Class Access': [2,2],
    'Application Access': [2,2],
    'Field Level Security (FLS)': [2,2],
    'Login IP Ranges': [2,2],
    'Object Level Security (OLS)': [2,2],
    'Record Type Creation': [2,2],
    'Page Layout Definitions': [2,2],
    'Tab Access': [2,2],
    'User Permission Access': [2,2],
    'Visualforce Page Access': [2,2],
    'Custom Metadata Type Access': [2,2],
    'Custom Setting Access': [2,2],
    'License': [1,1],
}


class Utils:
    def autoSizeColumns(workbook):
        for ws in workbook.sheetnames:
            if 'Sheet' in ws:
                workbook.remove(workbook[ws])
            else:
                for col in workbook[ws].columns:
                    max_length = 0
                    column = get_column_letter(
                        col[0].column)  # Get the column name
                    for cell in col:
                        try:  # Necessary to avoid error on empty cells
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    workbook[ws].column_dimensions[column].width = adjusted_width

    def processMDfile(workbook,file, templateDict):
        processedRows = 0
        #workbook = Workbook()
        sheet = None
        currentSheetName = None
        row = 1
        column = 1
        contentData = []
        profileName = ''

        for line in file.readlines():
            if line != '\n' and '===' not in line:
                # print(line)
                if processedRows == 0:
                    profileName = line
                else:
                    # Content header
                    if line.startswith('# '):
                        # remove square brackets
                        line = line[line.find('# ')+1:line.find(' (`')].strip()
                        contentData.append(line)
                        sheet = workbook.create_sheet(line)
                        currentSheetName = line
                        if type(templateDict.get(currentSheetName)) != set:
                            templateDict[currentSheetName] = set()
                        # reset to start on cell A1
                        row = 1
                        column = 1
                        # print(line)
                    else:
                        # Process content based on header
                        if sheet:
                            if not line.startswith('*'):
                                lineContent = line.split('|')
                                for content in lineContent:
                                    cell = sheet.cell(row, column)
                                    cell.value = content
                                    if row >= 2 and column == 2:
                                        if 'Field Level Security (FLS)' == currentSheetName or 'Record Type Creation' == currentSheetName or 'Page Layout Definitions' == currentSheetName:
                                            templateDict.get(currentSheetName).add(
                                                lineContent[1]+'|'+lineContent[2])
                                        elif 'Login IP Ranges' == currentSheetName:  # TODO fix array index of bound exception
                                            try:
                                                description = lineContent[1] or ' '
                                                start = lineContent[2] or ' '
                                                end = lineContent[3] or ' '
                                                templateDict.get(currentSheetName).add(
                                                    description+'|'+start+'|'+end)
                                            except:
                                                templateDict.get(currentSheetName).add(
                                                    lineContent[1])
                                        else:
                                            templateDict.get(currentSheetName).add(
                                                lineContent[1])
                                    elif row == 1 and column == 1 and (currentSheetName in ['License','User License','Description']):
                                        templateDict.get(currentSheetName).add(lineContent[0])
                                    column += 1
                                row += 1
                                column = 1
                processedRows = processedRows+1
        # remove square brackets
        profileName = profileName[profileName.find('[')+1:profileName.find(']')].strip()

        return profileName


def generateTemplate(templateDict, templateFileName):
    # Merge Template file creation
    templateWB = Workbook()
    row = 1
    column = 1
    for templateKey in templateDict.keys():
        # print(templateKey)
        ws = templateWB.create_sheet(templateKey)
        row = 1
        # print header
        for header in templateHeaders.get(templateKey):
            cell = ws.cell(row, column)
            cell.value = header
            column += 1
        row += 1
        column = 1
        for value in sorted(templateDict[templateKey]):
            # print(value)
            lineContent = value.split('|')
            for x in range(len(lineContent)):
                cell = ws.cell(row, x+1)
                cell.value = lineContent[x]
            row += 1

    Utils.autoSizeColumns(templateWB)
    # print(templateWB.sheetnames)
    templateWB.save(filename=urllib.parse.unquote(templateFileName, 'UTF-8'))


def mdToExcelwithTemplateGeneration(fileNameRegex, templateFileName,aliasName):
    os.chdir(path=os.curdir+'/input')
    templateDict = {}
    for fileName in glob.glob(fileNameRegex):
        # print(file)
        file = open(fileName)
        workbook = Workbook()

        # print(workbook.sheetnames)
        filename = Utils.processMDfile(workbook,file, templateDict)
        Utils.autoSizeColumns(workbook)
        workbook.save(filename=urllib.parse.unquote(filename, 'UTF-8')+aliasName)
    generateTemplate(templateDict, templateFileName)


def mdToTemplate(fileNameRegex, templateFileName, mergeFileName):
    os.chdir(path=os.curdir+'/input')
    wb2 = load_workbook(templateFileName)
    # print(wb2.sheetnames)
    templateDict = {}
    # Load template into map
    for ws in wb2.sheetnames:
        if ws not in templateDict.keys():
            templateDict[ws] = {}

        startIndex = templateStartRowColumn.get(ws)
        for value in wb2[ws].iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
            key = ''
            for txt in value:
                if None is not txt:
                    if '' != key:
                        key += '|'
                    key += txt

            if key not in templateDict[ws].keys():
                templateDict[ws][key] = {}

    # Load target files
    for fileName in glob.glob(fileNameRegex):
        tempWS = load_workbook(fileName)

        for sheetsKey in templateDict.keys():
            for itemKey in templateDict[sheetsKey]:
                if fileName not in templateDict[sheetsKey][itemKey].keys():
                    templateDict[sheetsKey][itemKey][fileName] = '--'

            if sheetsKey in tempWS.sheetnames:
                headerSize = templateHeaders.get(sheetsKey)
                startIndex = templateStartRowColumn.get(sheetsKey)
                for value in tempWS[sheetsKey].iter_rows(min_row=startIndex[0], min_col=startIndex[1], max_col=4, values_only=True):
                    key = ''
                    for x in range(len(headerSize)):
                        if value[x]:
                            key += value[x]
                        elif sheetsKey in ['License','User License','Description','Login IP Ranges'] and x == 0:
                                key += ' '
                        if x+1 < len(headerSize):
                            key += '|'
                    #TODO fix it better
                    if key.startswith('|'):
                        key = key[1:]

                    if key in templateDict[sheetsKey].keys():
                        if sheetsKey in ['License','User License','Description','Login IP Ranges']:
                            templateDict[sheetsKey][key][fileName] = '✔'
                        else:
                            templateDict[sheetsKey][key][fileName] = value[len(headerSize)]
        tempWS = None

    for sheets in templateDict.keys():
        sheet = wb2[sheets]
        row = 2
        column = 4
        for items in sorted(templateDict[sheets].keys()):
            column = 4
            for key in templateDict[sheets][items]:
                title = key.split('.')
                sheet.cell(1, column).value = title[0]
                value = templateDict[sheets][items].get(key)
                sheet.cell(row, column).value = value
                column += 1
            row += 1

    Utils.autoSizeColumns(wb2)
    wb2.save(filename=mergeFileName)

    # Load target MD files
    # for fileName in glob.glob(fileNameRegex):
    # print(file)
    #    file = open(fileName)
    #    workbook = None

    # print(workbook.sheetnames)
    #workbook = Utils.processMDfile(file, templateDict)
    # Utils.autoSizeColumns(workbook)
    # workbook.save(filename=file.name+".xlsx")


profileMDFileNameRegex = '*profile*md'
profileFileNameRegex = '*profile*md.xlsx'
profileTemplateFileName = 'profileTemplate.xlsx'
profileMergeFileName = 'profile-merged.xlsx'
profileAliasFile = '.profile.md.xlsx'

permissionsetMDFileNameRegex = '*permissionset*md'
permissionsetFileNameRegex = '*permissionset*md.xlsx'
permissionsetTemplateFileName = 'permissionTemplate.xlsx'
permissionMergeFileName = 'permissionset-merged.xlsx'
permissionAliasFile = '.permissionset.md.xlsx'

#TODO Create a more friendly logic with menu to select which option to execute
# Permission set template and excel generation
# 50 Permission Sets: 4s
#mdToExcelwithTemplateGeneration(permissionsetMDFileNameRegex,permissionsetTemplateFileName,permissionAliasFile)

# Merge all permission set excel files into one
# 50 Permission Sets: 14s
#mdToTemplate(permissionsetFileNameRegex,permissionsetTemplateFileName,permissionMergeFileName)

# Profile template and excel generation
# 150 files: 12m 23s
mdToExcelwithTemplateGeneration(profileMDFileNameRegex,profileTemplateFileName,profileAliasFile)

# Merge all profile excel files into one
# 150 files: 25m 56s
# mdToTemplate(profileFileNameRegex, profileTemplateFileName,profileMergeFileName)

